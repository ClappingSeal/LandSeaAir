from dronekit import Command, VehicleMode, connect, LocationGlobalRelative, APIException
from pymavlink.dialects.v20 import ardupilotmega
from pymavlink import mavutil
import time
import logging
import threading
from openpyxl import Workbook
import math

logging.getLogger('dronekit').setLevel(logging.CRITICAL)


class Drone:
    def __init__(self, connection_string='COM5', baudrate=57600):
        print('vehicle connecting...')

        # Connecting value
        self.connection_string = connection_string
        self.baudrate = baudrate
        self.vehicle = connect(self.connection_string, wait_ready=True, baud=self.baudrate, timeout=60)

        # Communication value
        self.vehicle.add_message_listener('DATA64', self.on_data64)
        self.received_data = None
        self.async_result = None

        # Position value
        self.init_lat = self.vehicle.location.global_relative_frame.lat
        self.init_lon = self.vehicle.location.global_relative_frame.lon

    # 통신 함수

    def on_data64(self, vehicle, name, message):
        if isinstance(message, ardupilotmega.MAVLink_data64_message):
            data = [int.from_bytes(message.data[i:i + 4], 'little') for i in range(0, len(message.data), 4)]
            self.received_data = data

    def get_received_data(self):
        return self.received_data

    def close_connection(self):
        self.vehicle.close()

    @staticmethod
    def asynchronous_received_data(drone_receiver):
        while True:
            if drone_receiver.get_received_data():
                a = drone_receiver.get_received_data()
                drone_receiver.async_result = a
                time.sleep(0.1)

    # 드론 액션 함수

    def set_flight_mode_by_pwm(self, pwm_value):
        def set_rc_channel_pwm(channel, pwm_value):
            if channel < 1:
                return
            self.vehicle.channels.overrides[channel] = pwm_value

        if 0 <= pwm_value <= 1230:
            mode = 'STABILIZE'
        elif 1231 <= pwm_value <= 1360:
            mode = 'AUTO'
        elif 1361 <= pwm_value <= 1490:
            mode = 'GUIDED'
        else:
            print("Another PWM value.")
            return

        set_rc_channel_pwm(5, pwm_value)
        print(f"mode to {mode}...")

    def arm(self):
        try:
            self.vehicle.channels.overrides['3'] = 1000
            self.set_flight_mode_by_pwm(1000)  # pwm signal for Stabilize mode
            time.sleep(3)
            self.vehicle.armed = True
            time.sleep(3)  # Wait for the drone to be armed
            if self.vehicle.armed:
                print("Vehicle armed")
            else:
                print("Vehicle armed fail")
        except APIException as e:
            print(str(e))

    def takeoff(self, h):
        self.set_flight_mode_by_pwm(1000)
        time.sleep(0.1)
        self.vehicle.channels.overrides['3'] = 1500
        time.sleep(2)

        cmds = self.vehicle.commands
        cmds.download()
        cmds.wait_ready()
        cmds.clear()
        takeoff_cmd = Command(0, 0, 0, mavutil.mavlink.MAV_FRAME_GLOBAL_RELATIVE_ALT,
                              mavutil.mavlink.MAV_CMD_NAV_TAKEOFF, 0, 0, 0, 0, 0, 0, 0, 0, h)
        cmds.add(takeoff_cmd)
        cmds.upload()

        self.set_flight_mode_by_pwm(1300)  # pwm signal for AUTO mode
        time.sleep(0.1)

        while not self.vehicle.armed:
            print("Waiting for arming...")
            time.sleep(1)

        print("Taking off!")

        # Monitoring altitude
        while True:
            print(f"Altitude: {self.vehicle.location.global_relative_frame.alt}")
            if self.vehicle.location.global_relative_frame.alt >= h * 0.8:
                print("Reached target altitude")
                break
            time.sleep(0.5)

    def north_direction(self):
        self.set_flight_mode_by_pwm(1400)  # pwm signal for GUIDED mode
        time.sleep(0.1)
        
        yaw_angle = 0
        is_relative = 0
        clockwise = 0
        yaw_rate = 0

        yaw_cmd = Command(0, 0, 0, 0,
                          mavutil.mavlink.MAV_CMD_CONDITION_YAW,
                          0, 0, yaw_angle, yaw_rate, clockwise, is_relative,
                          0, 0, 0)

        cmds = self.vehicle.commands
        cmds.clear()
        cmds.add(yaw_cmd)
        cmds.upload()

        self.set_flight_mode_by_pwm(1300)  # pwm signal for AUTO mode
        time.sleep(2)

    def goto_auto(self, x, y, z, velocity=4):
        LATITUDE_CONVERSION = 111000
        LONGITUDE_CONVERSION = 88.649 * 1000

        base_lat = self.init_lat
        base_lon = self.init_lon

        target_lat = base_lat + (y / LATITUDE_CONVERSION)
        target_lon = base_lon + (x / LONGITUDE_CONVERSION)

        self.set_flight_mode_by_pwm(1400)  # pwm signal for GUIDED mode
        time.sleep(0.1)

        cmds = self.vehicle.commands
        cmds.download()
        cmds.wait_ready()
        cmds.clear()

        goto_cmd = Command(0, 0, 0, mavutil.mavlink.MAV_FRAME_GLOBAL_RELATIVE_ALT,
                           mavutil.mavlink.MAV_CMD_NAV_WAYPOINT, 0, 0, 0, 0, velocity, 0, target_lat, target_lon, z)
        cmds.add(goto_cmd)
        cmds.upload()

        self.set_flight_mode_by_pwm(1300)  # pwm signal for AUTO mode

        # 모니터 링
        while math.dist([self.vehicle.location.global_relative_frame.lat,
                         self.vehicle.location.global_relative_frame.lon,
                         self.vehicle.location.global_relative_frame.alt],
                        [target_lat, target_lon, z]) > 0.5:  # Adjust as needed
            print(f"Current Position: {self.vehicle.location.global_relative_frame}")
            time.sleep(0.5)
        print("Reached target position")

    def goto_guided(self, x, y, z, velocity=4):
        LATITUDE_CONVERSION = 111000
        LONGITUDE_CONVERSION = 88.649 * 1000

        base_lat = self.init_lat
        base_lon = self.init_lon

        target_lat = base_lat + (y / LATITUDE_CONVERSION)
        target_lon = base_lon + (x / LONGITUDE_CONVERSION)

        self.set_flight_mode_by_pwm(1400)  # pwm signal for GUIDED mode
        time.sleep(0.1)

        # Set target location and velocity in GUIDED mode
        location = LocationGlobalRelative(target_lat, target_lon, z)
        self.vehicle.simple_goto(location, groundspeed=velocity)

        # Monitoring the position
        while math.dist([self.vehicle.location.global_relative_frame.lat,
                         self.vehicle.location.global_relative_frame.lon,
                         self.vehicle.location.global_relative_frame.alt],
                        [target_lat, target_lon, z]) > 0.5:  # Adjust as needed
            print(f"Current Position: {self.vehicle.location.global_relative_frame}")
            time.sleep(0.5)
        print("Reached target position")

    def land_by_auto_mode(self):
        self.set_flight_mode_by_pwm(1400)
        time.sleep(0.1)
        print("Landing using AUTO mode...")
        cmds = self.vehicle.commands
        cmds.download()
        cmds.wait_ready()
        cmds.clear()

        land_cmd = Command(0, 0, 0, mavutil.mavlink.MAV_FRAME_GLOBAL_RELATIVE_ALT,
                           mavutil.mavlink.MAV_CMD_NAV_LAND, 0, 0, 0, 0, 0, 0,
                           self.vehicle.location.global_relative_frame.lat,
                           self.vehicle.location.global_relative_frame.lon, 0)  # 0 for altitude as it's landing

        cmds.add(land_cmd)
        cmds.upload()

        self.vehicle.mode = VehicleMode("AUTO")
        self.set_flight_mode_by_pwm(1300)  # pwm signal for AUTO mode

        while self.vehicle.location.global_relative_frame.alt > 0.1:
            print(f"Altitude: {self.vehicle.location.global_relative_frame.alt}")
            time.sleep(1)

        print("Landed successfully!")

    # # RTL 고도/속도 조절 (추후에 반드시 해야함)

    def RTL(self, h=3, velocity=4):
        target_lat = self.init_lat
        target_lon = self.init_lon

        self.set_flight_mode_by_pwm(1400)  # pwm signal for GUIDED mode
        time.sleep(0.1)

        cmds = self.vehicle.commands
        cmds.download()
        cmds.wait_ready()
        cmds.clear()

        goto_cmd = Command(0, 0, 0, mavutil.mavlink.MAV_FRAME_GLOBAL_RELATIVE_ALT,
                           mavutil.mavlink.MAV_CMD_NAV_WAYPOINT, 0, 0, 0, 0, velocity, 0, target_lat, target_lon, h)
        land_cmd = Command(0, 0, 0, mavutil.mavlink.MAV_FRAME_GLOBAL_RELATIVE_ALT,
                           mavutil.mavlink.MAV_CMD_NAV_LAND, 0, 0, 0, 0, 0, 0,
                           self.vehicle.location.global_relative_frame.lat,
                           self.vehicle.location.global_relative_frame.lon, 0)  # 0 for altitude as it's landing
        cmds.add(goto_cmd)
        cmds.add(land_cmd)
        cmds.upload()

        self.set_flight_mode_by_pwm(1300)  # pwm signal for AUTO mode

        # 모니터 링
        while math.dist([self.vehicle.location.global_relative_frame.lat,
                         self.vehicle.location.global_relative_frame.lon,
                         self.vehicle.location.global_relative_frame.alt],
                        [target_lat, target_lon, h]) > 0.5:  # Adjust as needed
            print(f"Current Position: {self.vehicle.location.global_relative_frame}")
            time.sleep(0.5)
        print("Landed successfully!")

    # 그 외 함수들

    @staticmethod
    def position_excel(drone_receiver):
        # Create a new Excel workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Drone Positions"

        # Write the header to the Excel sheet
        headers = ['Time', 'Latitude', 'Longitude', 'Altitude']
        for col_num, header in enumerate(headers, 1):
            col_letter = ws.cell(row=1, column=col_num)
            col_letter.value = header

        row_number = 2
        while True:
            lat = drone_receiver.vehicle.location.global_relative_frame.lat
            lon = drone_receiver.vehicle.location.global_relative_frame.lon
            alt = drone_receiver.vehicle.location.global_relative_frame.alt

            # Write the new data to the next row in the Excel sheet
            ws.cell(row=row_number, column=1,
                    value=row_number - 1)  # Here we use row_number - 1 for sequential numbering
            ws.cell(row=row_number, column=2, value=lat)
            ws.cell(row=row_number, column=3, value=lon)
            ws.cell(row=row_number, column=4, value=alt)
            row_number += 1

            # Save the workbook to a file
            wb.save('drone_position.xlsx')

            time.sleep(1)


if __name__ == "__main__":
    # 드론 연결
    drone = Drone()

    # 데이터 수신 async 처리
    data_thread = threading.Thread(target=Drone.asynchronous_received_data, args=(drone,))
    data_thread.start()

    # 엑셀 저장 시작 (메인 프로젝트 시 제거)
    excel_thread = threading.Thread(target=Drone.position_excel, args=(drone,))
    excel_thread.start()

    # 드론 행동 블럭
    try:
        # 입력 (위도, 경도)
        raw_input = input("두 개의 실수를 입력하세요: ")
        nums = [float(num.strip()) for num in raw_input.split(",")]

        # 미션 시작
        if len(nums) == 2:
            print(drone.async_result)
            """
            drone.arm() # no delay
            drone.takeoff(3) # no delay
            drone.north_direction() # no delay
            drone.goto_auto(1, 2, 3, 3)
            time.sleep(3)
            drone.goto_auto(1, 10, 3, 3)
            time.sleep(3)
            drone.land_by_auto_mode()
            
            # drone.RTL()
            """


        else:
            print("정확하게 두 개의 실수를 입력하세요.")

    except ValueError:
        print("올바른 형식의 실수를 입력하세요.")
    except KeyboardInterrupt:
        drone.close_connection()

